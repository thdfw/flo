import os
import time
import numpy as np
import pandas as pd
import pendulum
import matplotlib
import matplotlib.pyplot as plt
from matplotlib.colors import Normalize, ListedColormap, BoundaryNorm
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
from utils import COP, to_celcius, get_data, required_SWT
from utils import (HORIZON_HOURS, MIN_TOP_TEMP_F, MAX_TOP_TEMP_F, TEMP_LIFT_F, NUM_LAYERS, 
                   MAX_HP_POWER_KW, MIN_HP_POWER_KW, STORAGE_VOLUME_GALLONS, STORAGE_LOSSES_PERCENT, 
                   START_TIME, INITIAL_TOP_TEMP_F, INITIAL_THERMOCLINE, SHOW_PLOT, NOW_FOR_FILE)

SOFT_CONSTRAINT = False

class Node():
    def __init__(self, time_slice:int, top_temp:float, thermocline:float):
        self.time_slice = time_slice
        self.top_temp = top_temp
        self.thermocline = thermocline
        self.energy = self.get_energy()
        self.pathcost = 0 if time_slice==HORIZON_HOURS else 1e9
        self.next_node = None
        self.index = None

    def __repr__(self):
        return f"Node[time_slice:{self.time_slice}, top_temp:{self.top_temp}, thermocline:{self.thermocline}]"

    def get_energy(self):
        m_layer_kg = STORAGE_VOLUME_GALLONS * 3.785 / NUM_LAYERS
        energy_above_kWh = self.thermocline*m_layer_kg * 4.187/3600 * to_celcius(self.top_temp)
        energy_below_kWh = (NUM_LAYERS-self.thermocline)*m_layer_kg * 4.187/3600 * to_celcius(self.top_temp - TEMP_LIFT_F)
        return energy_above_kWh + energy_below_kWh


class Edge():
    def __init__(self, tail:Node, head:Node, cost:float, heat_output_HP:float):
        self.tail = tail
        self.head = head
        self.cost = cost
        self.heat_output_HP = heat_output_HP

    def __repr__(self):
        return f"Edge: {self.tail} --cost:{round(self.cost,3)}--> {self.head}"


class Graph():
    def __init__(self, start_state:Node, start_time:pendulum.DateTime):
        print("\nSetting up the graph...")
        timer = time.time()
        self.start_time = start_time
        self.source_node = start_state
        self.get_forecasts()
        self.create_nodes()
        self.create_edges()
        print(f"Done in {round(time.time()-timer,2)} seconds.\n")

    def get_forecasts(self):
        df = get_data(self.start_time)
        self.elec_dist_prices = list(df.dist)
        self.elec_lmp = list(df.lmp)
        self.elec_prices = [x+y for x,y in zip(list(df.dist), list(df.lmp))]
        self.oat = list(df.oat)
        self.load = list(df.load)
        self.min_SWT = [required_SWT(x) for x in self.load]

    def create_nodes(self):
        self.nodes = {}
        for time_slice in range(HORIZON_HOURS+1):
            self.nodes[time_slice] = [self.source_node] if time_slice==0 else []
            self.nodes[time_slice].extend(
                Node(time_slice, top_temp, thermocline)
                for top_temp in range(MIN_TOP_TEMP_F, MAX_TOP_TEMP_F+TEMP_LIFT_F, TEMP_LIFT_F) 
                for thermocline in list(range(NUM_LAYERS+1))
                if (time_slice, top_temp, thermocline) != (0, self.source_node.top_temp, self.source_node.thermocline)
            )
            self.nodes_by_energy = sorted(self.nodes[time_slice], key=lambda x: (x.energy, x.top_temp), reverse=True)
            for n in self.nodes[time_slice]:
                n.index = self.nodes_by_energy.index(n)+1
    
    def create_edges(self):
        self.edges = {}
        self.min_energy_node = Node(0,MIN_TOP_TEMP_F,0).energy
        self.energy_between_consecutive_states = Node(0,MIN_TOP_TEMP_F,2).energy - Node(0,MIN_TOP_TEMP_F,1).energy
        for h in range(HORIZON_HOURS):
            for node_now in self.nodes[h]:
                self.edges[node_now] = []
                for node_next in self.nodes[h+1]:
                    heat_to_store = node_next.energy - node_now.energy
                    losses = STORAGE_LOSSES_PERCENT/100*(node_now.energy-self.min_energy_node)
                    if losses<self.energy_between_consecutive_states and losses>0 and self.load[h]==0:
                        losses = self.energy_between_consecutive_states + 1/1e9
                    heat_output_HP = heat_to_store + self.load[h] + losses
                    if heat_output_HP <= MAX_HP_POWER_KW and heat_output_HP >= MIN_HP_POWER_KW:
                        cop = COP(oat=self.oat[h], lwt=node_next.top_temp)
                        cost = self.elec_prices[h]/100 * heat_output_HP / cop
                        if heat_to_store > 0:
                            if node_next.top_temp==node_now.top_temp and node_next.thermocline>node_now.thermocline:
                                self.edges[node_now].append(Edge(node_now, node_next, cost, heat_output_HP))
                            if node_next.top_temp==node_now.top_temp+TEMP_LIFT_F:
                                self.edges[node_now].append(Edge(node_now, node_next, cost, heat_output_HP))
                        elif heat_to_store < 0:
                            if heat_output_HP < self.load[h]:
                                 if self.load[h]>0 and (node_now.top_temp < self.min_SWT[h] or node_next.top_temp < self.min_SWT[h]):
                                    if SOFT_CONSTRAINT:
                                         cost += 1e5
                                    else:
                                        continue
                            if node_next.top_temp==node_now.top_temp and node_next.thermocline<node_now.thermocline:
                                self.edges[node_now].append(Edge(node_now, node_next, cost, heat_output_HP))
                            if node_next.top_temp==node_now.top_temp-TEMP_LIFT_F:
                                self.edges[node_now].append(Edge(node_now, node_next, cost, heat_output_HP))
                        else:
                            self.edges[node_now].append(Edge(node_now, node_next, cost, heat_output_HP))

    def solve_dijkstra(self):
        print("Solving Dijkstra...")
        start_time = time.time()
        for time_slice in range(HORIZON_HOURS-1, -1, -1):
            for node in self.nodes[time_slice]:
                best_edge = min(self.edges[node], key=lambda e: e.head.pathcost + e.cost)
                if best_edge.heat_output_HP<0: 
                    best_edge_neg = max([e for e in self.edges[node] if e.heat_output_HP<0], key=lambda e: e.heat_output_HP)
                    best_edge_pos = min([e for e in self.edges[node] if e.heat_output_HP>=0], key=lambda e: e.heat_output_HP)
                    best_edge = best_edge_pos if (-best_edge_neg.heat_output_HP >= best_edge_pos.heat_output_HP) else best_edge_neg
                node.pathcost = best_edge.head.pathcost + best_edge.cost
                node.next_node = best_edge.head
        print(f"Done in {round(time.time()-start_time,3)} seconds.\n")

    def compute_bid(self):
        max_edge = max(self.edges[self.source_node], key=lambda x: x.heat_output_HP)
        min_edge = min(self.edges[self.source_node], key=lambda x: x.heat_output_HP)
        max_edge_elec = max_edge.cost/self.elec_prices[0]*100
        min_edge_elec = min_edge.cost/self.elec_prices[0]*100
        bid = (min_edge.head.pathcost - max_edge.head.pathcost) / (max_edge_elec - min_edge_elec)
        self.bid = bid
        print(f"Bid price: {round(bid*100,2)} cts/kWh\n")

    def plot(self):
        self.list_toptemps = []
        self.list_thermoclines = []
        self.list_hp_energy = []
        self.list_storage_energy = []
        node_i = self.source_node
        while node_i.next_node is not None:
            edge_i = [e for e in self.edges[node_i] if e.head==node_i.next_node][0]
            self.list_toptemps.append(node_i.top_temp)
            self.list_thermoclines.append(node_i.thermocline)
            self.list_hp_energy.append(edge_i.heat_output_HP)
            self.list_storage_energy.append(node_i.energy)
            node_i = node_i.next_node
        self.list_toptemps.append(node_i.top_temp)
        self.list_thermoclines.append(node_i.thermocline)
        self.list_hp_energy.append(edge_i.heat_output_HP)
        self.list_storage_energy.append(node_i.energy)
        max_energy_node = Node(0,MAX_TOP_TEMP_F,NUM_LAYERS).energy
        list_soc = [(x-self.min_energy_node)/(max_energy_node-self.min_energy_node)*100 for x in self.list_storage_energy]
        # Plot the shortest path
        fig, ax = plt.subplots(2,1, sharex=True, figsize=(10,6))
        begin = self.start_time.format('YYYY-MM-DD HH:mm')
        end = self.start_time.add(hours=HORIZON_HOURS).format('YYYY-MM-DD HH:mm')
        fig.suptitle(f'From {begin} to {end}\nCost: {round(self.source_node.pathcost,2)} $', fontsize=10)
        list_time = list(range(len(list_soc)))
        # Top plot
        ax[0].step(list_time, self.list_hp_energy, where='post', color='tab:blue', label='HP', alpha=0.6)
        ax[0].step(list_time, self.load, where='post', color='tab:red', label='Load', alpha=0.6)
        ax[0].legend(loc='upper left')
        ax[0].set_ylabel('Heat [kWh]')
        ax[0].set_ylim([-0.5,20])
        ax2 = ax[0].twinx()
        ax2.step(list_time, self.elec_prices, where='post', color='gray', alpha=0.6, label='Elec price')
        ax2.legend(loc='upper right')
        ax2.set_ylabel('Electricity price [cts/kWh]')
        if min(self.elec_prices)>0: ax2.set_ylim([0,60])
        # Bottom plot
        norm = Normalize(vmin=MIN_TOP_TEMP_F-TEMP_LIFT_F, vmax=MAX_TOP_TEMP_F)
        cmap = matplotlib.colormaps['Reds']
        tank_top_colors = [cmap(norm(x)) for x in self.list_toptemps]
        tank_bottom_colors = [cmap(norm(x-TEMP_LIFT_F)) for x in self.list_toptemps]
        list_thermoclines_reversed = [NUM_LAYERS-x+1 for x in self.list_thermoclines]
        ax[1].bar(list_time, list_thermoclines_reversed, color=tank_bottom_colors, alpha=0.7)
        ax[1].bar(list_time, self.list_thermoclines, bottom=list_thermoclines_reversed, color=tank_top_colors, alpha=0.7)
        ax[1].set_xlabel('Time [hours]')
        ax[1].set_ylabel('Storage state')
        ax[1].set_ylim([0, NUM_LAYERS])
        ax[1].set_yticks([])
        if len(list_time)>10 and len(list_time)<50:
            ax[1].set_xticks(list(range(0,len(list_time)+1,2)))
        ax3 = ax[1].twinx()
        ax3.plot(list_time, list_soc, color='black', alpha=0.4, label='SoC')
        ax3.set_ylabel('State of charge [%]')
        ax3.set_ylim([-1,101])
        # Color bar
        boundaries = np.arange(MIN_TOP_TEMP_F-TEMP_LIFT_F*1.5, MAX_TOP_TEMP_F+TEMP_LIFT_F*1.5, TEMP_LIFT_F)
        colors = [plt.cm.Reds(i/(len(boundaries)-1)) for i in range(len(boundaries))]
        cmap = ListedColormap(colors)
        norm = BoundaryNorm(boundaries, cmap.N, clip=True)
        sm = plt.cm.ScalarMappable(cmap=cmap, norm=norm)
        cbar = plt.colorbar(sm, ax=ax, orientation='horizontal', fraction=0.025, pad=0.15, alpha=0.7)
        cbar.set_ticks(range(MIN_TOP_TEMP_F-TEMP_LIFT_F, MAX_TOP_TEMP_F+TEMP_LIFT_F, TEMP_LIFT_F))
        cbar.set_label('Temperature [F]')
        plt.savefig('plot.png', dpi=130)
        if SHOW_PLOT:
            plt.show()

    def export_excel(self):
        self.plot()
        print("Exporting to Excel...")
        start_time = time.time()
        # Along the shortest path
        electricitiy_used, heat_delivered = [], []
        node_i = self.source_node
        while node_i.next_node is not None:
            heat_to_store = node_i.next_node.energy - node_i.energy
            losses = STORAGE_LOSSES_PERCENT/100*(node_i.energy-self.min_energy_node)
            if losses<self.energy_between_consecutive_states and losses>0 and self.load[node_i.time_slice]==0:
                losses = self.energy_between_consecutive_states + 1/1e9
            heat_output_HP = heat_to_store + self.load[node_i.time_slice] + losses
            cop = COP(oat=self.oat[node_i.time_slice], lwt=node_i.next_node.top_temp)
            electricitiy_used.append(heat_output_HP / cop)
            heat_delivered.append(heat_output_HP)
            node_i = node_i.next_node
        # First dataframe: the Dijkstra graph
        dijkstra_pathcosts = {}
        dijkstra_pathcosts['Top Temp [F]'] = [x.top_temp for x in self.nodes_by_energy]
        dijkstra_pathcosts['Thermocline'] = [x.thermocline for x in self.nodes_by_energy]
        dijkstra_pathcosts['Index'] = list(range(1,len(self.nodes_by_energy)+1))
        dijkstra_nextnodes = dijkstra_pathcosts.copy()
        for h in range(HORIZON_HOURS):
            dijkstra_pathcosts[h] = [round(x.pathcost,2) for x in sorted(self.nodes[h], key=lambda x: x.index)]
            dijkstra_nextnodes[h] = [x.next_node.index for x in sorted(self.nodes[h], key=lambda x: x.index)]
        dijkstra_pathcosts[HORIZON_HOURS] = [0 for x in self.nodes[HORIZON_HOURS]]
        dijkstra_nextnodes[HORIZON_HOURS] = [np.nan for x in self.nodes[HORIZON_HOURS]]
        dijkstra_pathcosts_df = pd.DataFrame(dijkstra_pathcosts)
        dijkstra_nextnodes_df = pd.DataFrame(dijkstra_nextnodes)
        # Second dataframe: the forecasts
        forecast_df = pd.DataFrame({'Forecast':['0'], 'Unit':['0'], **{h: [0.0] for h in range(HORIZON_HOURS+1)}})
        forecast_df.loc[0] = ['Price - total'] + ['cts/kWh'] + self.elec_prices
        forecast_df.loc[1] = ['Price - distribution'] + ['cts/kWh'] + self.elec_dist_prices
        forecast_df.loc[2] = ['Price - LMP'] + ['cts/kWh'] + self.elec_lmp
        forecast_df.loc[3] = ['Heating load'] + ['kW'] + [round(x,2) for x in self.load]
        forecast_df.loc[4] = ['OAT'] + ['F'] + [round(x,2) for x in self.oat]
        forecast_df.loc[5] = ['Required SWT'] + ['F'] + [round(x) for x in self.min_SWT]
        # Third dataframe: the shortest path
        shortestpath_df = pd.DataFrame({'Shortest path':['0'], 'Unit':['0'], **{h: [0.0] for h in range(HORIZON_HOURS+1)}})
        shortestpath_df.loc[0] = ['Electricity used'] + ['kWh'] + [round(x,3) for x in electricitiy_used] + [0]
        shortestpath_df.loc[1] = ['Heat delivered'] + ['kWh'] + [round(x,3) for x in heat_delivered] + [0]
        shortestpath_df.loc[2] = ['Cost - total'] + ['cts'] + [round(x*y,2) for x,y in zip(electricitiy_used, self.elec_prices)] + [0]
        shortestpath_df.loc[3] = ['Cost - distribution'] + ['cts'] + [round(x*y,2) for x,y in zip(electricitiy_used, self.elec_dist_prices)] + [0]
        shortestpath_df.loc[4] = ['Cost - LMP'] + ['cts'] + [round(x*y,2) for x,y in zip(electricitiy_used, self.elec_lmp)] + [0]
        # Final dataframe: the results
        total_usd = round(self.source_node.pathcost,2)
        total_elec = round(sum(electricitiy_used),2)
        total_heat = round(sum(heat_delivered),2)
        next_index = self.source_node.next_node.index
        results = ['Cost ($)', total_usd, 'Electricity (kWh)', total_elec, 'Heat (kWh)', total_heat, 'Next step index', next_index]
        results_df = pd.DataFrame({'RESULTS':results})
        # Highlight shortest path
        highlight_positions = []
        node_i = self.source_node
        while node_i.next_node is not None:
            highlight_positions.append((node_i.index+len(forecast_df)+len(shortestpath_df)+2, 3+node_i.time_slice))
            node_i = node_i.next_node
        highlight_positions.append((node_i.index+len(forecast_df)+len(shortestpath_df)+2, 3+node_i.time_slice))
        # Read the parameters
        parameters = {}
        with open('parameters.conf', 'r') as file:
            for line in file:
                stripped_line = line.strip()
                if stripped_line and not stripped_line.startswith('#'):
                    key_value = stripped_line.split('=')
                    if len(key_value) == 2:
                        key = key_value[0].strip()
                        value = key_value[1].strip()
                        parameters[key] = value
        parameters_df = pd.DataFrame(list(parameters.items()), columns=['Variable', 'Value'])
        # Write to Excel
        os.makedirs('results', exist_ok=True)
        file_path = os.path.join('results', f'result_{NOW_FOR_FILE}.xlsx')
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            results_df.to_excel(writer, index=False, sheet_name='Pathcost')
            results_df.to_excel(writer, index=False, sheet_name='Next node')
            forecast_df.to_excel(writer, index=False, startcol=1, sheet_name='Pathcost')
            forecast_df.to_excel(writer, index=False, startcol=1, sheet_name='Next node')
            shortestpath_df.to_excel(writer, index=False, startcol=1, startrow=len(forecast_df)+1, sheet_name='Pathcost')
            shortestpath_df.to_excel(writer, index=False, startcol=1, startrow=len(forecast_df)+1, sheet_name='Next node')
            dijkstra_pathcosts_df.to_excel(writer, index=False, startrow=len(forecast_df)+len(shortestpath_df)+2, sheet_name='Pathcost')
            dijkstra_nextnodes_df.to_excel(writer, index=False, startrow=len(forecast_df)+len(shortestpath_df)+2, sheet_name='Next node')
            parameters_df.to_excel(writer, index=False, sheet_name='Parameters')
            pathcost_sheet = writer.sheets['Pathcost']
            nextnode_sheet = writer.sheets['Next node']
            parameters_sheet = writer.sheets['Parameters']
            plot_sheet = writer.book.create_sheet(title='Plot')
            plot_sheet.add_image(Image('plot.png'), 'A1')
            for row in pathcost_sheet['A1:A10']:
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(bold=True)
            for row in nextnode_sheet['A1:A10']:
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(bold=True)
            pathcost_sheet.column_dimensions['A'].width = 15
            pathcost_sheet.column_dimensions['B'].width = 15
            pathcost_sheet.column_dimensions['C'].width = 15
            nextnode_sheet.column_dimensions['A'].width = 15
            nextnode_sheet.column_dimensions['B'].width = 15
            nextnode_sheet.column_dimensions['C'].width = 15
            parameters_sheet.column_dimensions['A'].width = 40
            parameters_sheet.column_dimensions['B'].width = 70
            pathcost_sheet.freeze_panes = 'D14'
            nextnode_sheet.freeze_panes = 'D14'
            highlight_fill = PatternFill(start_color='72ba93', end_color='72ba93', fill_type='solid')
            for row in range(len(forecast_df)+len(shortestpath_df)+2):
                pathcost_sheet.cell(row=row+1, column=1).fill = highlight_fill
                nextnode_sheet.cell(row=row+1, column=1).fill = highlight_fill
            for row, col in highlight_positions:
                pathcost_sheet.cell(row=row+1, column=col+1).fill = highlight_fill
                nextnode_sheet.cell(row=row+1, column=col+1).fill = highlight_fill
        os.remove('plot.png')
        print(f"Done in {round(time.time()-start_time,2)} seconds.\n")


if __name__ == '__main__':
    
    time_now = START_TIME
    state_now = Node(time_slice=0, top_temp=INITIAL_TOP_TEMP_F, thermocline=INITIAL_THERMOCLINE)

    g = Graph(state_now, time_now)
    g.solve_dijkstra()
    g.export_excel()
    # g.compute_bid()
    g.plot()